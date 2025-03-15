from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, current_app, send_file
from flask_login import login_required, current_user
from ..models.models import Event, Booking, db
from datetime import datetime, timezone
from sqlalchemy import text
from ..utils.email import send_event_registration_confirmation, send_admin_registration_notification
import traceback
import io
from openpyxl import Workbook

bp = Blueprint('main', __name__)

@bp.route('/')
def index():
    current_app.logger.info("Accessing index route")
    current_app.logger.info(f"User authenticated: {current_user.is_authenticated}")
    if current_user.is_authenticated:
        current_app.logger.info(f"User is admin: {current_user.is_admin}")

    # Debug: Check all events in database
    all_events = Event.query.all()
    current_app.logger.info(f"Total events in database: {len(all_events)}")
    for event in all_events:
        current_app.logger.info(f"Event in DB: {event.title}, Date: {event.date}, Visible: {event.is_visible}")

    if current_user.is_authenticated and current_user.is_admin:
        # Admin sees all future events, including invisible ones
        events = Event.get_future_events(include_invisible=True)
        current_app.logger.info(f"Admin view - Future events count: {len(events)}")
        for event in events:
            current_app.logger.info(f"Admin future event: {event.title}, Date: {event.date}")
        return render_template('index.html', events=events)
    else:
        # Non-admin users only see visible future events
        events = Event.get_future_events(include_invisible=False)
        current_app.logger.info(f"User view - Future events count: {len(events)}")
        for event in events:
            current_app.logger.info(f"User future event: {event.title}, Date: {event.date}")
        return render_template('index.html', events=events)

@bp.route('/event/create', methods=['GET', 'POST'])
@login_required
def create_event():
    if request.method == 'POST':
        try:
            title = request.form['title']
            description = request.form['description']
            date_str = request.form['date']
            capacity = int(request.form.get('capacity', 10))
            room = request.form.get('room')
            address = request.form.get('address')
            price = float(request.form.get('price', 0))
            
            # Parse the date and make it timezone-aware
            date = datetime.strptime(date_str, '%Y-%m-%dT%H:%M').replace(tzinfo=timezone.utc)
            
            event = Event(
                title=title,
                description=description,
                date=date,
                capacity=capacity,
                room=room,
                address=address,
                price=price
            )
            db.session.add(event)
            db.session.commit()
            
            flash('Veranstaltung erfolgreich erstellt!', 'success')
            return redirect(url_for('main.index'))
        except ValueError as e:
            db.session.rollback()
            current_app.logger.error(f"Fehler bei der Erstellung der Veranstaltung: {str(e)}\n{traceback.format_exc()}")
            flash(str(e), 'danger')
            return render_template('create_event.html', 
                                default_date=date_str,
                                form_data=request.form)
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Fehler bei der Erstellung der Veranstaltung: {str(e)}\n{traceback.format_exc()}")
            flash('Ein Fehler ist aufgetreten, während die Veranstaltung erstellt wurde.', 'danger')
            return render_template('create_event.html', 
                                default_date=date_str,
                                form_data=request.form)
    
    default_date = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M')
    return render_template('create_event.html', default_date=default_date)

@bp.route('/event/<int:event_id>/book', methods=['GET', 'POST'])
def book_event(event_id):
    """Book an event."""
    event = Event.query.get_or_404(event_id)
    
    if request.method == 'POST':
        if event.bookings >= event.capacity:
            flash('Diese Veranstaltung ist leider ausgebucht!', 'error')
            return redirect(url_for('main.index'))
        
        name = request.form['name']
        email = request.form['email']
        phone = request.form['phone']
        
        booking = Booking(event_id=event_id, name=name, email=email, phone=phone)
        
        try:
            # Start transaction
            db.session.begin_nested()  # Create a savepoint
            
            # Add booking and increment counter
            db.session.add(booking)
            event.bookings += 1
            
            # Try to send confirmation email to user first
            send_event_registration_confirmation(email, event)
            
            # Send notification to admin
            user_data = {'name': name, 'email': email, 'phone': phone}
            send_admin_registration_notification(event, user_data)
            
            # If emails sent successfully, commit the transaction
            db.session.commit()
            flash('Buchung erfolgreich! Wir haben Sie für die Veranstaltung gebucht.', 'success')
            
            #if current_app.config.get('DISABLE_EMAILS', False):
            #    flash('Buchung erfolgreich! Eine Bestätigungs-E-Mail wurde an Ihre E-Mail-Adresse gesendet.', 'success')
            #else:
            #    flash('Buchung erfolgreich! Wir haben Sie für die Veranstaltung gebucht.', 'success')

            return redirect(url_for('main.book_event', event_id=event_id))
            
        except Exception as e:
            # Roll back to the savepoint
            db.session.rollback()
            current_app.logger.error(f"Fehler bei der Buchung: {str(e)}\n{traceback.format_exc()}")
            flash('Bei der Verarbeitung Ihrer Buchung ist ein Fehler aufgetreten. Bitte versuchen Sie es erneut.', 'error')
            return redirect(url_for('main.book_event', event_id=event_id))
    
    return render_template('book_event.html', event=event)

@bp.route('/event/<int:event_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_event(event_id):
    # Use get() instead of get_or_404() for better performance
    event = Event.query.get(event_id)
    if not event:
        flash('Veranstaltung nicht gefunden.', 'danger')
        return redirect(url_for('main.index'))
    
    if request.method == 'POST':
        try:
            # Batch update attributes
            form_data = {
                'title': request.form['title'],
                'description': request.form['description'],
                'capacity': int(request.form.get('capacity', 10)),
                'room': request.form.get('room'),
                'address': request.form.get('address'),
                'price': float(request.form.get('price', 0))
            }
            
            # Update all fields at once
            for key, value in form_data.items():
                setattr(event, key, value)
            
            # Handle date separately due to timezone
            date_str = request.form['date']
            event.date = datetime.strptime(date_str, '%Y-%m-%dT%H:%M').replace(tzinfo=timezone.utc)
            
            # Use a single commit for all changes
            db.session.commit()
            
            flash('Veranstaltung erfolgreich aktualisiert!', 'success')
            return redirect(url_for('main.index'))
            
        except ValueError as e:
            db.session.rollback()
            current_app.logger.error(f"Fehler bei der Aktualisierung der Veranstaltung: {str(e)}\n{traceback.format_exc()}")
            flash(str(e), 'danger')
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Fehler bei der Aktualisierung der Veranstaltung: {str(e)}\n{traceback.format_exc()}")
            flash('Ein Fehler ist aufgetreten, während die Veranstaltung aktualisiert wurde.', 'danger')
    
    return render_template('edit_event.html', event=event)

@bp.route('/event/<int:event_id>/copy', methods=['POST'])
@login_required
def copy_event(event_id):
    try:
        # Get the original event
        original_event = Event.query.get_or_404(event_id)
        
        # Ensure the date is timezone-aware
        event_date = original_event.date
        if event_date.tzinfo is None:
            event_date = event_date.replace(tzinfo=timezone.utc)
        
        # Create a new event with the same details
        new_event = Event(
            title=f"Kopie von {original_event.title}",
            description=original_event.description,
            date=event_date,
            capacity=original_event.capacity,
            room=original_event.room,
            address=original_event.address,
            price=original_event.price,
            bookings=0  # Start with 0 bookings
        )
        
        db.session.add(new_event)
        db.session.commit()
        
        flash('Veranstaltung erfolgreich kopiert!', 'success')
        return redirect(url_for('main.index'))
    except ValueError as e:
        db.session.rollback()
        current_app.logger.error(f"Fehler bei der Kopie der Veranstaltung: {str(e)}\n{traceback.format_exc()}")
        flash(str(e), 'danger')
        return redirect(url_for('main.index'))
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Fehler bei der Kopie der Veranstaltung: {str(e)}\n{traceback.format_exc()}")
        flash('Ein Fehler ist aufgetreten, während die Veranstaltung kopiert wurde.', 'danger')
        return redirect(url_for('main.index'))

@bp.route('/event/<int:event_id>/toggle-visibility', methods=['POST'])
@login_required
def toggle_visibility(event_id):
    try:
        event = Event.query.get_or_404(event_id)
        event.is_visible = not event.is_visible
        db.session.commit()
        status = "sichtbar" if event.is_visible else "unsichtbar"
        flash(f'Veranstaltung ist jetzt {status}', 'success')
        return redirect(url_for('main.index'))
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Fehler bei der Aktualisierung der Sichtbarkeit der Veranstaltung: {str(e)}\n{traceback.format_exc()}")
        flash('Fehler bei der Aktualisierung der Sichtbarkeit der Veranstaltung', 'danger')
        return redirect(url_for('main.index'))

@bp.route('/event/<int:event_id>/registrations')
@login_required
def view_registrations(event_id):
    event = Event.query.get_or_404(event_id)
    bookings = Booking.query.filter_by(event_id=event_id).order_by(Booking.created_at.desc()).all()
    return render_template('registrations.html', event=event, bookings=bookings)

@bp.route('/event/<int:event_id>/export')
@login_required
def export_registrations(event_id):
    """Export event registrations as Excel file."""
    # Ensure user is admin
    if not current_user.is_admin:
        flash('Zugriff verweigert. Sie benötigen Administratorrechte.', 'danger')
        return redirect(url_for('main.index'))
    
    event = Event.query.get_or_404(event_id)
    bookings = Booking.query.filter_by(event_id=event_id).order_by(Booking.created_at.desc()).all()
    
    # Create a workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Registrierungen"
    
    # Add headers
    headers = ["Name", "Telefonnummer", "E-Mail"]
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    # Add data
    for row_num, booking in enumerate(bookings, 2):
        worksheet.cell(row=row_num, column=1, value=booking.name)
        worksheet.cell(row=row_num, column=2, value=booking.phone)
        worksheet.cell(row=row_num, column=3, value=booking.email)
    
    # Create a BytesIO object to store the Excel file
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    # Generate a filename with the event date in YYYY-MM-DD format
    event_date_str = event.date.strftime('%Y-%m-%d')
    filename = f"{event_date_str}-Anmeldungen.xlsx"
    
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/booking/<int:booking_id>/delete', methods=['POST'])
@login_required
def delete_booking(booking_id):
    booking = Booking.query.get_or_404(booking_id)
    event = Event.query.get(booking.event_id)
    
    # Decrement the bookings count
    event.bookings = max(0, event.bookings - 1)
    
    # Delete the booking
    try:
        db.session.delete(booking)
        db.session.commit()
        flash('Anmeldung erfolgreich gelöscht.', 'success')
        return redirect(url_for('main.view_registrations', event_id=booking.event_id))
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Fehler bei der Löschung der Anmeldung: {str(e)}\n{traceback.format_exc()}")
        flash('Ein Fehler ist aufgetreten, während die Anmeldung gelöscht wurde.', 'error')
        return redirect(url_for('main.view_registrations', event_id=booking.event_id))

@bp.route('/event/<int:event_id>/delete', methods=['POST'])
@login_required
def delete_event(event_id):
    if not current_user.is_admin:
        flash('Sie haben keine Berechtigung, Veranstaltungen zu löschen.', 'error')
        return redirect(url_for('main.index'))
    
    event = Event.query.get_or_404(event_id)
    try:
        # Delete associated bookings first
        Booking.query.filter_by(event_id=event_id).delete()
        
        # Delete the event
        db.session.delete(event)
        db.session.commit()
        
        flash('Veranstaltung erfolgreich gelöscht.', 'success')
        return redirect(url_for('main.index'))
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Fehler bei der Löschung der Veranstaltung: {str(e)}\n{traceback.format_exc()}")
        flash('Ein Fehler ist aufgetreten, während die Veranstaltung gelöscht wurde.', 'error')
        return redirect(url_for('main.index'))

@bp.route('/health')
def health_check():
    """Health check endpoint for Docker container."""
    try:
        # Check database connection
        db.session.execute(text('SELECT 1'))
        
        # Return success response
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'timestamp': datetime.now().isoformat()
        }), 200
    except Exception as e:
        current_app.logger.error(f"Health check failed: {str(e)}")
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@bp.route('/admin/check-reminders')
@login_required
def check_reminders():
    """Manually trigger reminder check (admin only)."""
    if not current_user.is_admin:
        flash('Sie haben keine Berechtigung für diese Aktion.', 'danger')
        return redirect(url_for('main.index'))
    
    try:
        from ..utils.scheduler import schedule_event_reminders
        schedule_event_reminders(current_app._get_current_object())
        flash('Erinnerungsprüfung wurde erfolgreich ausgeführt.', 'success')
    except Exception as e:
        current_app.logger.error(f"Error checking reminders: {str(e)}")
        flash(f'Fehler bei der Erinnerungsprüfung: {str(e)}', 'danger')
    
    return redirect(url_for('main.index'))

@bp.route('/admin/reminders/status')
@login_required
def reminder_status():
    """Get the current status of the reminder system (admin only)."""
    if not current_user.is_admin:
        flash('Sie haben keine Berechtigung für diese Aktion.', 'danger')
        return redirect(url_for('main.index'))
    
    try:
        from ..utils.scheduler import get_scheduler_status
        status = get_scheduler_status()
        return render_template('reminder_status.html', status=status)
    except Exception as e:
        current_app.logger.error(f"Error getting reminder status: {str(e)}")
        flash(f'Fehler beim Abrufen des Erinnerungsstatus: {str(e)}', 'danger')
        return redirect(url_for('main.index'))

@bp.route('/admin/reminders/toggle', methods=['POST'])
@login_required
def toggle_reminders():
    """Toggle the reminder system on/off (admin only)."""
    if not current_user.is_admin:
        flash('Sie haben keine Berechtigung für diese Aktion.', 'danger')
        return redirect(url_for('main.index'))
    
    try:
        # Get current setting
        current_setting = current_app.config.get('ENABLE_REMINDERS', True)
        
        # Toggle setting
        new_setting = not current_setting
        current_app.config['ENABLE_REMINDERS'] = new_setting
        
        # Update environment variable for persistence across restarts
        # Note: This doesn't modify the .env file, just the current process environment
        import os
        os.environ['ENABLE_REMINDERS'] = str(new_setting)
        
        status = "aktiviert" if new_setting else "deaktiviert"
        flash(f'Erinnerungen wurden {status}.', 'success')
    except Exception as e:
        current_app.logger.error(f"Error toggling reminders: {str(e)}")
        flash(f'Fehler beim Umschalten der Erinnerungen: {str(e)}', 'danger')
    
    return redirect(url_for('main.reminder_status'))
